import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date
from tools.db import get_all_requests

EXPORT_PATH = Path("data/requests_export.xlsx")

STATUS_FR = {
    "SUBMITTED":   "Soumis",
    "ASSIGNED":    "Assigné",
    "IN PROGRESS": "En cours",
    "DONE":        "Terminé",
    "CANCELLED":   "Annulé",
}

COLUMNS = [
    ("ID",              "id",               10),
    ("Date soumission", "date_submitted",   16),
    ("Secteur",         "secteur",          10),
    ("Type",            "type",             18),
    ("Demandeur",       "requester",        22),
    ("Description",     "description",      50),
    ("Statut",          "status",           14),
    ("Assigné à",       "assignee",         22),
    ("Date assigné",    "date_assigned",    16),
    ("Date en cours",   "date_in_progress", 16),
    ("Date terminé",    "date_done",        16),
    ("Délai (jours)",   "aging_days",       14),
    ("Pièce jointe",    "attachment_path",  30),
]

HEADER_FILL   = PatternFill("solid", fgColor="1A56DB")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL      = PatternFill("solid", fgColor="F1F5F9")
BORDER_SIDE   = Side(style="thin", color="E2E8F0")
CELL_BORDER   = Border(bottom=Border(bottom=BORDER_SIDE).bottom)

STATUS_FILLS = {
    "SUBMITTED":   PatternFill("solid", fgColor="E5E7EB"),
    "ASSIGNED":    PatternFill("solid", fgColor="DBEAFE"),
    "IN PROGRESS": PatternFill("solid", fgColor="FEF3C7"),
    "DONE":        PatternFill("solid", fgColor="D1FAE5"),
    "CANCELLED":   PatternFill("solid", fgColor="FCE7F3"),
}


def generate_export() -> Path:
    requests = get_all_requests()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demandes"

    # Header row
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, req in enumerate(requests, start=2):
        alt = row_idx % 2 == 0

        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            value = req.get(field)

            # Format special fields
            if field == "secteur" and value:
                value = f"Secteur {value}"
            elif field == "status" and value:
                value = STATUS_FR.get(value, value)
            elif field == "aging_days" and value is None:
                value = ""
            elif field == "attachment_path" and value:
                value = value.replace("uploads/", "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "description"))

            # Status column coloring
            if field == "status":
                raw_status = req.get("status", "")
                cell.fill = STATUS_FILLS.get(raw_status, PatternFill())
            elif alt:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 40 if req.get("description", "") else 18

    # Summary sheet
    ws2 = wb.create_sheet("Résumé")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12

    ws2.cell(1, 1, "Rapport généré le").font = Font(bold=True)
    ws2.cell(1, 2, date.today().isoformat())

    ws2.cell(3, 1, "Statut").font = Font(bold=True, color="FFFFFF")
    ws2.cell(3, 1).fill = HEADER_FILL
    ws2.cell(3, 2, "Nombre").font = Font(bold=True, color="FFFFFF")
    ws2.cell(3, 2).fill = HEADER_FILL

    from collections import Counter
    counts = Counter(r["status"] for r in requests)
    for i, (status, count) in enumerate(counts.items(), start=4):
        ws2.cell(i, 1, STATUS_FR.get(status, status))
        ws2.cell(i, 2, count)
        if status in STATUS_FILLS:
            ws2.cell(i, 1).fill = STATUS_FILLS[status]
            ws2.cell(i, 2).fill = STATUS_FILLS[status]

    ws2.cell(4 + len(counts), 1, "TOTAL").font = Font(bold=True)
    ws2.cell(4 + len(counts), 2, len(requests)).font = Font(bold=True)

    EXPORT_PATH.parent.mkdir(exist_ok=True)
    wb.save(EXPORT_PATH)
    return EXPORT_PATH
