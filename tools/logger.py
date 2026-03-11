"""
Append-only Excel audit log for all request events.
File: data/requests_log.xlsx

Each row = one event (creation or field change).
Never overwrites existing rows — always appends.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

LOG_PATH = Path("data/requests_log.xlsx")

HEADERS = [
    "Horodatage",
    "Type",
    "ID Demande",
    "Secteur",
    "Demandeur",
    "Champ modifié",
    "Ancienne valeur",
    "Nouvelle valeur",
    "Description (actuelle)",
    "Statut (actuel)",
    "Assigné à (actuel)",
]

HEADER_FILL = PatternFill("solid", fgColor="1A56DB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

EVENT_FILLS = {
    "Création":     PatternFill("solid", fgColor="D1FAE5"),
    "Modification": PatternFill("solid", fgColor="DBEAFE"),
}

COL_WIDTHS = [20, 14, 12, 10, 20, 18, 25, 25, 40, 14, 20]

STATUS_FR = {
    "SUBMITTED":   "Soumis",
    "ASSIGNED":    "Assigné",
    "IN PROGRESS": "En cours",
    "DONE":        "Terminé",
    "CANCELLED":   "Annulé",
}


def _ensure_workbook():
    """Return (workbook, worksheet), creating the file + header if needed."""
    LOG_PATH.parent.mkdir(exist_ok=True)

    if LOG_PATH.exists():
        wb = openpyxl.load_workbook(LOG_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal"

        for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        wb.save(LOG_PATH)

    return wb, ws


def _append_row(event_type, request_id, secteur, requester,
                field_changed, old_value, new_value,
                current_description, current_status, current_assignee, current_type=None):
    wb, ws = _ensure_workbook()

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        event_type,
        request_id,
        f"Secteur {secteur}" if secteur else "",
        requester or "",
        field_changed or "",
        old_value or "",
        new_value or "",
        current_description or "",
        STATUS_FR.get(current_status, current_status or ""),
        current_assignee or "",
    ]

    next_row = ws.max_row + 1
    fill = EVENT_FILLS.get(event_type, PatternFill())

    for col, value in enumerate(row, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=(col == 9))
        if col <= 2:
            cell.fill = fill

    ws.row_dimensions[next_row].height = 18
    wb.save(LOG_PATH)


def log_creation(request_id, secteur, requester, description, type=None):
    """Log a new request submission."""
    _append_row(
        event_type="Création",
        request_id=request_id,
        secteur=secteur,
        requester=requester,
        field_changed="—",
        old_value="—",
        new_value="—",
        current_description=description,
        current_status="SUBMITTED",
        current_assignee=None,
        current_type=type,
    )


def log_changes(request_id, old_record, new_assignee, new_status, new_description, new_type=None):
    """Log each changed field as a separate row."""
    changes = []

    if new_description is not None and new_description != old_record.get("description"):
        changes.append((
            "Description",
            old_record.get("description", ""),
            new_description,
        ))

    if new_assignee is not None and new_assignee != old_record.get("assignee"):
        changes.append((
            "Assigné à",
            old_record.get("assignee") or "—",
            new_assignee or "—",
        ))

    if new_type is not None and new_type != old_record.get("type"):
        changes.append((
            "Type",
            old_record.get("type") or "—",
            new_type or "—",
        ))

    if new_status is not None and new_status != old_record.get("status"):
        changes.append((
            "Statut",
            STATUS_FR.get(old_record.get("status", ""), old_record.get("status", "")),
            STATUS_FR.get(new_status, new_status),
        ))

    # Resolve what the final state will be (for context columns)
    final_status      = new_status      if new_status      is not None else old_record.get("status")
    final_assignee    = new_assignee    if new_assignee    is not None else old_record.get("assignee")
    final_description = new_description if new_description is not None else old_record.get("description")
    final_type        = new_type        if new_type        is not None else old_record.get("type")

    for field, old_val, new_val in changes:
        _append_row(
            event_type="Modification",
            request_id=request_id,
            secteur=old_record.get("secteur"),
            requester=old_record.get("requester"),
            field_changed=field,
            old_value=old_val,
            new_value=new_val,
            current_description=final_description,
            current_status=final_status,
            current_assignee=final_assignee,
            current_type=final_type,
        )
