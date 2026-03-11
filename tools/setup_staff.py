"""
Setup script: creates data/configuration.xlsx with sample Requesters and Engineers.

Usage:
    python tools/setup_staff.py          # create (skip if exists)
    python tools/setup_staff.py --force  # overwrite

To add/remove people: open data/configuration.xlsx in Excel, edit rows, save, restart server.
"""
import sys
import openpyxl
from pathlib import Path

STAFF_FILE = Path("data/configuration.xlsx")

SAMPLE_REQUESTERS = [
    "Alice Martin",
    "Bob Dupont",
    "Claire Leblanc",
    "David Moreau",
    "Emma Bernard",
]

SAMPLE_TYPES = [
    "Amélioration",
    "Documentation",
    "Formation",
    "Installation",
    "Inspection",
    "Maintenance",
    "Réparation",
    "Sécurité",
]

SAMPLE_ENGINEERS = [
    "Jean Technicien",
    "Marie Ingenieur",
    "Pierre Expert",
    "Sophie Analyste",
]


def create_staff_file(force=False):
    Path("data").mkdir(exist_ok=True)
    if STAFF_FILE.exists() and not force:
        print(f"[skip] {STAFF_FILE} already exists. Use --force to recreate.")
        return

    wb = openpyxl.Workbook()
    ws_req = wb.active
    ws_req.title = "Requesters"
    ws_req.append(["Name"])
    for name in SAMPLE_REQUESTERS:
        ws_req.append([name])

    ws_eng = wb.create_sheet("Engineers")
    ws_eng.append(["Name"])
    for name in SAMPLE_ENGINEERS:
        ws_eng.append([name])

    ws_types = wb.create_sheet("Types")
    ws_types.append(["Type"])
    for t in SAMPLE_TYPES:
        ws_types.append([t])

    wb.save(STAFF_FILE)
    print(f"[ok] Created {STAFF_FILE} with {len(SAMPLE_REQUESTERS)} requesters, {len(SAMPLE_ENGINEERS)} engineers and {len(SAMPLE_TYPES)} types.")


if __name__ == "__main__":
    create_staff_file(force="--force" in sys.argv)
