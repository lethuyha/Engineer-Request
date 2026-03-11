import openpyxl
from pathlib import Path

CONFIG_FILE = Path("data/configuration.xlsx")

DEFAULTS = {
    "Requesters": sorted([
        "Marc Dupont", "Lucie Bernard", "Jean Moreau",
        "Marie Lefebvre", "Paul Girard", "Claire Rousseau",
    ]),
    "Engineers": sorted([
        "Sophie Martin", "Thomas Leroy", "Ahmed Benali",
    ]),
    "Types": sorted([
        "Amélioration process", "Modification technique", "Analyse qualité",
        "Qualification", "Étude faisabilité", "Documentation", "Maintenance préventive",
    ]),
    "Secteurs": sorted([
        "Fonderie", "Usinage", "Qualité", "HSE", "Maintenance",
    ]),
}


def _read_sheet(sheet_name):
    """Read column A (from row 2) of a named sheet. Falls back to defaults if file missing."""
    if not CONFIG_FILE.exists():
        return DEFAULTS.get(sheet_name, [])
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return DEFAULTS.get(sheet_name, [])
    ws = wb[sheet_name]
    values = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    wb.close()
    return sorted(values) if values else DEFAULTS.get(sheet_name, [])


def get_staff():
    return {
        "requesters": _read_sheet("Requesters"),
        "engineers":  _read_sheet("Engineers"),
        "types":      _read_sheet("Types"),
        "secteurs":   _read_sheet("Secteurs"),
    }
